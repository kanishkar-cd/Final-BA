from __future__ import annotations

import logging
from pathlib import Path
from app.parsers.base_parser import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel spreadsheet files (.xlsx, .xls)."""

    async def parse(self, file_path: str) -> str:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel document not found: {path}")

        extracted_rows: list[str] = []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(filename=str(path), data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                extracted_rows.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    row_vals = [str(val).strip() for val in row if val is not None and str(val).strip()]
                    if row_vals:
                        extracted_rows.append(" | ".join(row_vals))
            return "\n".join(extracted_rows)
        except Exception as exc:
            logger.warning("openpyxl parsing failed for %s, using fallback reader: %s", path, exc)
            try:
                # Fallback CSV/text reading
                text_content = path.read_text(encoding="utf-8", errors="ignore")
                return text_content if text_content.strip() else f"Excel File: {path.name}"
            except Exception:
                return f"Excel File: {path.name}"
